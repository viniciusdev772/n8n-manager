"""
Parser do Relatório de Saldo de Abastecimento (fab0257)
Extrai itens (falta), substitutos, cores e valor devido (Abast negativo).

Uso:
    python parse_falta.py <caminho_do_pdf>

Saída:
    JSON → <nome>_parsed.json
    CSV  → <nome>_parsed.csv

Estrutura de colunas no PDF:
    Item original  (falta)     : x0 ≈ 14.4  | cor em x0 ≈ 164.2
    Item substituto (importado): x0 ≈ 17.6  | cor em x0 ≈ 156.4
    Unid (M/MT)                : x0 ≈ 149
    Dados numéricos (ignorar)  : x0 ≥ 226
    Abast (valor devido)       : x0 ≈ 245–285
"""

import sys, re, json, os, tempfile, csv
from datetime import UTC, datetime
from pathlib import Path
from urllib.parse import quote
from typing import Any, Dict, List, Optional
from html import escape
import pdfplumber
import pandas as pd
from collections import defaultdict
try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.formatting.rule import ColorScaleRule
    OPENPYXL_AVAILABLE = True
except ModuleNotFoundError:
    Workbook = None
    Alignment = None
    Border = None
    Font = None
    PatternFill = None
    Side = None
    get_column_letter = None
    Table = None
    TableStyleInfo = None
    ColorScaleRule = None
    OPENPYXL_AVAILABLE = False
try:
    from fastapi import FastAPI, File, HTTPException, Request, UploadFile
    from fastapi.openapi.utils import get_openapi
    from fastapi.responses import FileResponse
    from pydantic import BaseModel, Field
    FASTAPI_AVAILABLE = True
except ModuleNotFoundError:
    FastAPI = None
    File = None
    HTTPException = Exception
    Request = None
    UploadFile = None
    FileResponse = None
    get_openapi = None
    BaseModel = object
    Field = None
    FASTAPI_AVAILABLE = False

# ── Thresholds de coluna ─────────────────────────────────────────────────────
X_ORIGINAL_MAX   = 16.0   # item original começa em ~14.4  (x0 < 16)
X_SUBSTITUTO_MAX = 20.0   # substituto começa em ~17.6     (16 ≤ x0 < 20)
X_ITEM_DESC_MAX  = 145    # descrição do item termina antes de 145

X_COLOR_ORIG_MIN = 149    # inclui casos com código de cor "PAR..." em x0~149
X_COLOR_SUB_MIN  = 153    # cor do substituto começa ~156
X_COLOR_MAX      = 226    # dados numéricos começam ~230 (ignorar)

X_ABAST_MIN      = 240    # coluna Abast começa ~249
X_ABAST_MAX      = 290    # coluna Abast termina antes de ~290
X_TAM_MIN        = 226    # coluna Tam fica logo antes de Abast
X_TAM_MAX        = 240
X_SALDO_CASA_MIN = 330    # coluna "em casa" começa ~335
X_SALDO_CASA_MAX = 370    # coluna "em casa" termina ~364

# ── Comportamento ─────────────────────────────────────────────────────────────
INCLUDE_SUBSTITUTES = False  # True = inclui importados/substitutos no output
INCLUDE_PAR_SINGLETONS_IN_COMMON = True  # True = inclui PAR no common mesmo em 1 mini

SKIP_PATTERNS = re.compile(
    r"(Empresa:|Filial:|Usuário:|Tipo\s+Relatório|Grupo\s+de|Grupo\s+POI|"
    r"fab0257|Relatório\s+de|Página|Descrição|Unid\.Cor|Mini\s+Fabrica|"
    r"Total\s+de|CALCADOS|BEIRA\s+RIO|S/A|FILIAL|Somente\s+Falta|"
    r"\d{2}/\d{2}/\d{4}|\d{2}:\d{2}:\d{2}|Usuário:)",
    re.IGNORECASE,
)

ITEM_CODE_RE  = re.compile(r"^\d{4,6}$")
COLOR_CODE_RE = re.compile(r"^(?:\d{1,6}|PAR\d{4,})$")
ABAST_RE      = re.compile(r"^-[\d,\.]+$")
NUMERIC_RE    = re.compile(r"^-?[\d,]+\.\d+$|^-?[\d,]+$")
TAM_RE        = re.compile(r"^\d{1,2}$")
MINI_FABRICA_RE = re.compile(r"Mini\s+Fabrica\s*-\s*(.+)$", re.IGNORECASE)


def clean_color_word(word):
    """
    Remove o '0' do Tam que vaza para o último token da cor.
    Ocorre quando x1 do token ultrapassa X_COLOR_MAX (226) e
    o PDF fundiu o último char da cor com o '0' do Tam.
    """
    if word["x1"] > X_COLOR_MAX and word["text"].endswith("0"):
        return word["text"][:-1]
    return word["text"]


def strip_unid_bleed(item_zone):
    """
    Remove o 'M' ou 'MT' da coluna Unid que vaza para o último token
    da descrição do item quando a descrição é longa (x1 > 148).
    """
    words = item_zone[2:]  # pula código e traço
    if not words:
        return ""
    texts = [w["text"] for w in words]
    last_word = words[-1]
    if last_word["x1"] > 148:
        t = texts[-1]
        if t.endswith("MT"):
            texts[-1] = t[:-2]
        elif t.endswith("M"):
            texts[-1] = t[:-1]
    return re.sub(r"^-\s*", "", " ".join(texts)).strip()


def group_rows(words, y_tol=2.0):
    rows = defaultdict(list)
    for w in words:
        key = round(w["top"] / y_tol) * y_tol
        rows[key].append(w)
    return dict(sorted(rows.items()))


def is_header_row(row_words):
    full = " ".join(w["text"] for w in row_words)
    return bool(SKIP_PATTERNS.search(full))


def extract_color(color_zone):
    """Extrai código e descrição de cor a partir da zona de cor."""
    if not color_zone:
        return None, None, ""

    # Procura o padrão "<codigo> - <descricao>" em qualquer posição da zona.
    # Isso evita falha quando há token de unidade ("M"/"MT") antes do código.
    for i in range(len(color_zone) - 1):
        first = color_zone[i]["text"]
        second = color_zone[i + 1]["text"]
        if COLOR_CODE_RE.match(first) and second == "-":
            prev_is_par = i > 0 and color_zone[i - 1]["text"] == "PAR"
            par_tipo = "PAR" if first.startswith("PAR") or prev_is_par else ""
            code = first[3:] if first.startswith("PAR") else first
            desc = " ".join(clean_color_word(w) for w in color_zone[i + 2:]).strip()
            return code, desc, par_tipo
        # Alguns PDFs trazem PAR em token separado: "PAR 98071 - BRANCO"
        if (
            first == "PAR"
            and i + 2 < len(color_zone)
            and re.fullmatch(r"\d{1,6}", second)
            and color_zone[i + 2]["text"] == "-"
        ):
            code = second
            desc = " ".join(clean_color_word(w) for w in color_zone[i + 3:]).strip()
            return code, desc, "PAR"
    return None, None, ""


def extract_abast(row_words):
    """Extrai o valor de Abast (negativo) da linha, se existir."""
    abast_zone = [w for w in row_words if X_ABAST_MIN <= w["x0"] < X_ABAST_MAX]
    for w in abast_zone:
        if ABAST_RE.match(w["text"]):
            return float(w["text"].replace(",", ""))
    return None


def extract_saldo_casa(row_words):
    """Extrai o saldo em casa da linha, se existir."""
    casa_zone = [w for w in row_words if X_SALDO_CASA_MIN <= w["x0"] < X_SALDO_CASA_MAX]
    for w in casa_zone:
        if NUMERIC_RE.match(w["text"]):
            # Mantém as casas decimais do PDF (ex.: 177.00000).
            return w["text"].replace(",", "")
    return None


def extract_tam(row_words):
    """Extrai o valor de Tam da linha, quando existir."""
    tam_zone = [w for w in row_words if X_TAM_MIN <= w["x0"] < X_TAM_MAX]
    for w in tam_zone:
        if TAM_RE.match(w["text"]):
            return w["text"]
    return ""


def is_zero_saldo(value):
    if value is None or value == "":
        return True
    try:
        return float(str(value).replace(",", "")) == 0.0
    except ValueError:
        return False


def _has_value(value):
    return value is not None and value != ""


def _detail_quality(abast, saldo_casa, source_text):
    score = 0
    if abast is not None:
        score += 2
    if _has_value(saldo_casa):
        score += 2
    if str(source_text or "").strip():
        score += 1
    return score


def _choose_best_color_detail(existing: Dict[str, Any], candidate: Dict[str, Any]) -> Dict[str, Any]:
    """
    Escolhe o melhor registro entre dois candidatos da mesma cor.
    Critério principal: completude de dados (abast/saldo/texto origem).
    Empate: mantém o existente (ordem original do PDF).
    """
    existing_score = _detail_quality(
        existing.get("abast"),
        existing.get("saldo_casa"),
        existing.get("source_text"),
    )
    candidate_score = _detail_quality(
        candidate.get("abast"),
        candidate.get("saldo_casa"),
        candidate.get("source_text"),
    )
    if candidate_score > existing_score:
        chosen = dict(candidate)
    else:
        chosen = dict(existing)

    # Mantém descrições mais completas, independentemente do vencedor.
    existing_color_desc = str(existing.get("color_desc") or "")
    candidate_color_desc = str(candidate.get("color_desc") or "")
    if len(candidate_color_desc) > len(existing_color_desc):
        chosen["color_desc"] = candidate_color_desc
    else:
        chosen["color_desc"] = existing_color_desc

    return chosen


def _merge_destination_detail(existing: Optional[Dict[str, Any]], candidate: Dict[str, Any]) -> Dict[str, Any]:
    """
    Evita sobrescrever detalhe bom com linha parcial de quebra de página.
    """
    if existing is None:
        return dict(candidate)

    existing_score = _detail_quality(
        existing.get("abast"),
        existing.get("saldo_casa"),
        existing.get("source_text"),
    )
    candidate_score = _detail_quality(
        candidate.get("abast"),
        candidate.get("saldo_casa"),
        candidate.get("source_text"),
    )

    if candidate_score > existing_score:
        merged = dict(candidate)
        if not merged.get("saldo_origem"):
            merged["saldo_origem"] = existing.get("saldo_origem", "nacional")
    else:
        merged = dict(existing)
        if not merged.get("saldo_origem"):
            merged["saldo_origem"] = candidate.get("saldo_origem", "nacional")

    existing_source_text = str(existing.get("source_text") or "")
    candidate_source_text = str(candidate.get("source_text") or "")
    if len(candidate_source_text) > len(existing_source_text):
        merged["source_text"] = candidate_source_text
    elif "source_text" not in merged:
        merged["source_text"] = existing_source_text

    return merged


def merge_items_for_page_breaks(items):
    """
    Consolida duplicidades de item/cor na mesma mini fábrica.
    Resolve casos de quebra de página em que a primeira linha da página
    repete item/cor sem colunas numéricas.
    """
    merged_items: Dict[tuple, Dict[str, Any]] = {}
    order: List[tuple] = []

    for it in items:
        mini = it.get("mini_fabrica") or "Mini Fabrica - N/D"
        item_code = it.get("item_code", "")
        item_key = (mini, item_code)

        if item_key not in merged_items:
            merged_items[item_key] = {
                "item_code": item_code,
                "item_desc": it.get("item_desc", "") or "",
                "mini_fabrica": mini,
                "colors": [],
                "_color_index": {},
            }
            order.append(item_key)
        else:
            current_desc = merged_items[item_key].get("item_desc", "") or ""
            new_desc = it.get("item_desc", "") or ""
            if len(new_desc) > len(current_desc):
                merged_items[item_key]["item_desc"] = new_desc

        target = merged_items[item_key]
        for color in it.get("colors", []):
            color_key = (
                color.get("color_code", ""),
                color.get("par_tipo", "") or "",
                color.get("tam", "") or "",
            )
            idx = target["_color_index"].get(color_key)
            if idx is None:
                color_copy = dict(color)
                if color_copy.get("par_tipo") is None:
                    color_copy["par_tipo"] = ""
                if color_copy.get("tam") is None:
                    color_copy["tam"] = ""
                target["_color_index"][color_key] = len(target["colors"])
                target["colors"].append(color_copy)
            else:
                existing = target["colors"][idx]
                candidate = dict(color)
                if candidate.get("par_tipo") is None:
                    candidate["par_tipo"] = ""
                if candidate.get("tam") is None:
                    candidate["tam"] = ""
                target["colors"][idx] = _choose_best_color_detail(existing, candidate)

    result = []
    for key in order:
        item = merged_items[key]
        item["item_desc"] = (item.get("item_desc") or "").strip()
        del item["_color_index"]
        result.append(item)
    return result


def make_source_meta(page_number, row_words, row_text, color_zone=None):
    """Metadados de origem da linha no PDF para auditoria no HTML."""
    zone = color_zone or row_words
    first = zone[0] if zone else (row_words[0] if row_words else None)
    if first is None:
        return {
            "source_page": page_number,
            "source_x": None,
            "source_y": None,
            "source_text": row_text,
        }
    return {
        "source_page": page_number,
        "source_x": round(float(first.get("x0", 0.0)), 1),
        "source_y": round(float(first.get("top", 0.0)), 1),
        "source_text": row_text,
    }


def row_bbox(row_words):
    if not row_words:
        return None
    x0 = min(float(w.get("x0", 0.0)) for w in row_words)
    x1 = max(float(w.get("x1", x0)) for w in row_words)
    top = min(float(w.get("top", 0.0)) for w in row_words)
    bottom = max(float(w.get("bottom", top)) for w in row_words)
    return {"x0": x0, "x1": x1, "top": top, "bottom": bottom}


def _debug_enabled():
    return os.getenv("PARSER_EXPORT_DEBUG_IMAGES", "").strip().lower() in {"1", "true", "yes", "on"}


def save_extraction_debug_images(pdf, annotations_by_page: Dict[int, List[Dict[str, Any]]], pdf_path: str):
    base = os.path.splitext(pdf_path)[0]
    out_dir = Path(base + "_common_items_debug")
    out_dir.mkdir(parents=True, exist_ok=True)

    for page_idx, page in enumerate(pdf.pages, start=1):
        ann = annotations_by_page.get(page_idx, [])
        if not ann:
            continue
        try:
            img = page.to_image(resolution=140)
            item_rects = [a["bbox"] for a in ann if a.get("kind") == "item" and a.get("bbox")]
            color_rects = [a["bbox"] for a in ann if a.get("kind") == "color" and a.get("bbox")]
            extra_rects = [a["bbox"] for a in ann if a.get("kind") == "extra" and a.get("bbox")]

            if item_rects:
                img.draw_rects(item_rects, stroke="blue", stroke_width=2)
            if color_rects:
                img.draw_rects(color_rects, stroke="red", stroke_width=2)
            if extra_rects:
                img.draw_rects(extra_rects, stroke="orange", stroke_width=2)

            out_file = out_dir / f"page_{page_idx:03d}.png"
            img.save(str(out_file), format="PNG")
        except Exception as e:
            print(f"[WARN] Debug image P{page_idx} não gerada: {e}")

    print(f"[OK] Debug Imagens → {out_dir}")


def parse_pdf(pdf_path):
    """
    Retorna lista de itens originais (falta), cada um com:
      - item_code, item_desc
      - colors: [{color_code, color_desc, abast}]
      - substitutes: [{item_code, item_desc, colors: [{color_code, color_desc, abast}]}]
    """
    items        = []
    current_item = None
    current_sub  = None
    current_mini_fabrica = None

    with pdfplumber.open(pdf_path) as pdf:
        annotations_by_page: Dict[int, List[Dict[str, Any]]] = defaultdict(list)
        for page_number, page in enumerate(pdf.pages, start=1):
            # x_tolerance=1 reduz fusão indevida entre colunas vizinhas
            # (ex.: última letra da descrição + "M/MT" da coluna Unid).
            words = page.extract_words(x_tolerance=1, y_tolerance=2)
            rows  = group_rows(words)
            first_data_row_checked = False

            for _y, row_words in rows.items():
                row_words = sorted(row_words, key=lambda w: w["x0"])
                row_text = " ".join(w["text"] for w in row_words).strip()

                # Captura contexto da mini fabrica vigente no relatorio.
                mini_match = MINI_FABRICA_RE.search(row_text)
                if mini_match:
                    current_mini_fabrica = f"Mini Fabrica - {mini_match.group(1).strip()}"
                    continue

                if is_header_row(row_words):
                    continue

                first_x = row_words[0]["x0"] if row_words else 999

                item_zone  = [w for w in row_words if w["x0"] < X_ITEM_DESC_MAX]
                color_orig = [w for w in row_words if X_COLOR_ORIG_MIN <= w["x0"] < X_COLOR_MAX]
                color_sub  = [w for w in row_words if X_COLOR_SUB_MIN  <= w["x0"] < X_COLOR_MAX]

                first_item_word = item_zone[0]["text"] if item_zone else ""
                is_new_code     = ITEM_CODE_RE.match(first_item_word)

                abast = extract_abast(row_words)
                saldo_casa = extract_saldo_casa(row_words)
                tam = extract_tam(row_words)

                if not first_data_row_checked:
                    first_data_row_checked = True
                    # Guarda de continuidade entre páginas:
                    # se a página começa com continuação mas não há item ativo,
                    # essa linha fica órfã e precisa de revisão manual.
                    if not is_new_code and first_x < X_ORIGINAL_MAX and current_item is None:
                        print(
                            f"[WARN] P{page_number}: início com continuação sem item ativo: "
                            f"{row_text[:140]}"
                        )

                # ── ITEM ORIGINAL ─────────────────────────────────────────────
                if is_new_code and first_x < X_ORIGINAL_MAX:
                    item_desc = strip_unid_bleed(item_zone)
                    current_item = {
                        "item_code": first_item_word,
                        "item_desc": item_desc,
                        "mini_fabrica": current_mini_fabrica,
                        "colors":    [],
                        "_sub_saldo_by_color": {},
                    }
                    current_sub = False  # reset: não estamos em substituto
                    items.append(current_item)
                    item_box = row_bbox(row_words)
                    if item_box:
                        annotations_by_page[page_number].append(
                            {"kind": "item", "bbox": item_box, "item_code": first_item_word}
                        )

                    code, desc, par_tipo = extract_color(color_orig)
                    if code:
                        src = make_source_meta(page_number, row_words, row_text, color_orig)
                        current_item["colors"].append({
                            "color_code": code,
                            "color_desc": desc,
                            "par_tipo": par_tipo,
                            "tam": tam if par_tipo else "",
                            "abast":      abast,
                            "saldo_casa": saldo_casa,
                            "saldo_origem": "nacional",
                            **src,
                        })
                        color_box = row_bbox(color_orig or row_words)
                        if color_box:
                            annotations_by_page[page_number].append(
                                {"kind": "color", "bbox": color_box, "item_code": first_item_word, "color_code": code}
                            )

                # ── SUBSTITUTO — ignorar, só marcar flag ──────────────────────
                elif is_new_code and X_ORIGINAL_MAX <= first_x < X_SUBSTITUTO_MAX:
                    current_sub = True  # linhas seguintes de cor pertencem ao sub → ignorar
                    # Guarda saldo do substituto por cor para usar como fallback
                    # quando o saldo do item nacional estiver zerado.
                    if current_item:
                        sub_code, _sub_desc, _sub_par_tipo = extract_color(color_sub)
                        sub_saldo_casa = extract_saldo_casa(row_words)
                        if sub_code and sub_saldo_casa is not None:
                            current_item["_sub_saldo_by_color"][sub_code] = sub_saldo_casa

                # ── CONTINUAÇÃO DE DESCRIÇÃO OU COR EXTRA ────────────────────
                else:
                    if item_zone and current_item and not is_new_code and first_x < X_ORIGINAL_MAX:
                        current_item["item_desc"] += " " + " ".join(w["text"] for w in item_zone)

                    # Cor extra de item original.
                    # Mesmo após linha de substituto, ainda podem existir linhas de cor
                    # do item original (ex.: começam com "M" e cor em x0~164).
                    if not is_new_code and color_orig and current_item:
                        code, desc, par_tipo = extract_color(color_orig)
                        if code:
                            src = make_source_meta(page_number, row_words, row_text, color_orig)
                            current_item["colors"].append({
                                "color_code": code,
                                "color_desc": desc,
                                "par_tipo": par_tipo,
                                "tam": tam if par_tipo else "",
                                "abast":      abast,
                                "saldo_casa": saldo_casa,
                                "saldo_origem": "nacional",
                                **src,
                            })
                            color_box = row_bbox(color_orig or row_words)
                            if color_box:
                                annotations_by_page[page_number].append(
                                    {
                                        "kind": "color",
                                        "bbox": color_box,
                                        "item_code": current_item.get("item_code"),
                                        "color_code": code,
                                    }
                                )
                        elif current_item["colors"]:
                            extra = " ".join(clean_color_word(w) for w in color_orig).strip()
                            current_item["colors"][-1]["color_desc"] += " " + extra
                            extra_box = row_bbox(color_orig)
                            if extra_box:
                                annotations_by_page[page_number].append(
                                    {
                                        "kind": "extra",
                                        "bbox": extra_box,
                                        "item_code": current_item.get("item_code"),
                                        "color_code": current_item["colors"][-1].get("color_code"),
                                    }
                                )

        if _debug_enabled():
            save_extraction_debug_images(pdf, annotations_by_page, pdf_path)

    # Limpa espaços
    for it in items:
        sub_saldo_map = it.get("_sub_saldo_by_color", {})
        it["item_desc"] = it["item_desc"].strip()
        for c in it["colors"]:
            c["color_desc"] = c["color_desc"].strip()
            if is_zero_saldo(c.get("saldo_casa")) and c["color_code"] in sub_saldo_map:
                c["saldo_casa"] = sub_saldo_map[c["color_code"]]
                c["saldo_origem"] = "substituto"
            elif c.get("saldo_origem") is None:
                c["saldo_origem"] = "nacional"
            if c.get("par_tipo") is None:
                c["par_tipo"] = ""
            if c.get("tam") is None:
                c["tam"] = ""
        if "_sub_saldo_by_color" in it:
            del it["_sub_saldo_by_color"]

    return merge_items_for_page_breaks(items)


def save_json(items, path):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(items, f, ensure_ascii=False, indent=2)
    print(f"[OK] JSON → {path}")


def _build_common_distribution(common_items):
    def normalized_need(abast):
        if abast is None:
            return 0.0
        need = abs(float(abast))
        # Regra operacional: toda necessidade positiva abaixo de 1m
        # deve ser tratada como 1m mínimo.
        if 0.0 < need < 1.0:
            return 1.0
        return need

    distribution = []
    for it in common_items:
        destinos = []
        total_deve = 0.0
        for mf in it.get("mini_fabricas", []):
            det = it.get("por_mini_fabrica", {}).get(mf, {})
            abast = det.get("abast")
            need = normalized_need(abast)
            total_deve += need
            destinos.append(
                {
                    "mini_fabrica": mf,
                    "deve_abast": abast,
                    "necessidade": need,
                    "saldo_casa": det.get("saldo_casa", ""),
                    "saldo_origem": det.get("saldo_origem", ""),
                    "source_page": det.get("source_page"),
                    "source_x": det.get("source_x"),
                    "source_y": det.get("source_y"),
                    "source_text": det.get("source_text", ""),
                }
            )
        distribution.append(
            {
                "item_code": it.get("item_code", ""),
                "item_desc": it.get("item_desc", ""),
                "tipo_unidade": it.get("par_tipo", ""),
                "codigo_cor": it.get("color_code", ""),
                "descricao_cor": it.get("color_desc", ""),
                "tam": it.get("tam", ""),
                "qtd_mini_fabricas": len(destinos),
                "total_necessidade": round(total_deve, 4),
                "mini_fabricas_destino": [d["mini_fabrica"] for d in destinos],
                "destinos": destinos,
            }
        )
    return distribution


def _to_float(value):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        try:
            return float(value)
        except (TypeError, ValueError):
            return None
    text = str(value).strip().replace(" ", "")
    if not text:
        return None
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def _sanitize_excel_formula(value):
    if not isinstance(value, str):
        return value
    text = value
    if text.startswith(("=", "+", "@")):
        return "'" + text
    return text


def _prepare_dataframe_for_excel(df: pd.DataFrame, numeric_columns: List[str]) -> pd.DataFrame:
    out = df.copy()
    for col in numeric_columns:
        if col in out.columns:
            out[col] = out[col].apply(_to_float)
    for col in out.columns:
        if pd.api.types.is_object_dtype(out[col]):
            out[col] = out[col].apply(_sanitize_excel_formula)
    return out


def _save_styled_xlsx_from_df(
    df: pd.DataFrame,
    xlsx_path: str,
    sheet_name: str,
    table_name: str,
    numeric_formats: Optional[Dict[str, str]] = None,
    color_scale_columns: Optional[List[str]] = None,
) -> bool:
    if not OPENPYXL_AVAILABLE:
        print("[WARN] openpyxl não disponível: exportação XLSX não gerada.")
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = (sheet_name or "Dados")[:31]

    headers = list(df.columns)
    if not headers:
        wb.save(xlsx_path)
        return True

    ws.append(headers)
    for row in df.itertuples(index=False, name=None):
        ws.append([None if pd.isna(v) else v for v in row])

    max_row = ws.max_row
    max_col = ws.max_column

    header_fill = PatternFill(fill_type="solid", start_color="1F4E78", end_color="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(border_style="thin", color="D9E1F2")
    body_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    body_alt_fill = PatternFill(fill_type="solid", start_color="F8FBFF", end_color="F8FBFF")

    for col_idx in range(1, max_col + 1):
        cell = ws.cell(1, col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = body_border

    for row_idx in range(2, max_row + 1):
        use_alt = (row_idx % 2) == 0
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row_idx, col_idx)
            cell.border = body_border
            cell.alignment = body_alignment
            if use_alt:
                cell.fill = body_alt_fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

    safe_table_name = re.sub(r"[^A-Za-z0-9_]", "", table_name or "TableData")
    if not safe_table_name:
        safe_table_name = "TableData"
    if safe_table_name[0].isdigit():
        safe_table_name = "T_" + safe_table_name
    if max_row >= 2:
        tab = Table(displayName=safe_table_name[:80], ref=f"A1:{get_column_letter(max_col)}{max_row}")
        tab.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(tab)

    numeric_formats = numeric_formats or {}
    for col_name, fmt in numeric_formats.items():
        if col_name not in headers:
            continue
        col_idx = headers.index(col_name) + 1
        for row_idx in range(2, max_row + 1):
            cell = ws.cell(row_idx, col_idx)
            if isinstance(cell.value, (int, float)):
                cell.number_format = fmt
                cell.alignment = Alignment(horizontal="right", vertical="top")

    for col_name in color_scale_columns or []:
        if col_name not in headers or max_row < 2:
            continue
        col_idx = headers.index(col_name) + 1
        col_letter = get_column_letter(col_idx)
        ws.conditional_formatting.add(
            f"{col_letter}2:{col_letter}{max_row}",
            ColorScaleRule(
                start_type="min",
                start_color="F8696B",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFEB84",
                end_type="max",
                end_color="63BE7B",
            ),
        )

    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        sample_size = min(max_row, 800)
        max_len = len(str(header))
        for row_idx in range(2, sample_size + 1):
            value = ws.cell(row_idx, col_idx).value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 52)

    wb.save(xlsx_path)
    return True


def save_csv(items, path, xlsx_path=None):
    columns = [
        "Mini Fabrica", "Codigo Item", "Descricao Item", "Tipo Unidade",
        "Codigo Cor", "Descricao Cor", "Tam",
        "Deve (Abast)", "Saldo em Casa", "Origem do Saldo"
    ]
    records = []
    for it in items:
        for c in it["colors"]:
            records.append({
                "Mini Fabrica": it.get("mini_fabrica", ""),
                "Codigo Item": it["item_code"],
                "Descricao Item": it["item_desc"],
                "Tipo Unidade": c.get("par_tipo", ""),
                "Codigo Cor": c["color_code"],
                "Descricao Cor": c["color_desc"],
                "Tam": c.get("tam", ""),
                "Deve (Abast)": c.get("abast", ""),
                "Saldo em Casa": c.get("saldo_casa", ""),
                "Origem do Saldo": c.get("saldo_origem", ""),
            })

    # Mantém ordem de aparição do PDF (itens e cores), sem reordenação por código.
    df = pd.DataFrame.from_records(records, columns=columns)
    df = _prepare_dataframe_for_excel(df, numeric_columns=["Deve (Abast)", "Saldo em Casa"])
    target = Path(path)
    if target.suffix.lower() == ".xlsx":
        xlsx_ok = _save_styled_xlsx_from_df(
            df,
            xlsx_path=str(target),
            sheet_name="Itens",
            table_name="ItensParser",
            numeric_formats={
                "Deve (Abast)": "#,##0.00000",
                "Saldo em Casa": "#,##0.00000",
            },
            color_scale_columns=["Deve (Abast)", "Saldo em Casa"],
        )
        if xlsx_ok:
            print(f"[OK] XLSX → {target}")
        else:
            fallback_csv = str(target.with_suffix(".csv"))
            df.to_csv(
                fallback_csv,
                index=False,
                encoding="utf-8-sig",
                sep=";",
                decimal=",",
                quoting=csv.QUOTE_MINIMAL,
            )
            print(f"[WARN] XLSX indisponível, CSV fallback → {fallback_csv}")
        return

    # CSV amigável para Excel (pt-BR): delimitador ; e decimal com vírgula.
    df.to_csv(path, index=False, encoding="utf-8-sig", sep=";", decimal=",", quoting=csv.QUOTE_MINIMAL)
    print(f"[OK] CSV  → {path}")
    if xlsx_path:
        xlsx_ok = _save_styled_xlsx_from_df(
            df,
            xlsx_path=xlsx_path,
            sheet_name="Itens",
            table_name="ItensParser",
            numeric_formats={
                "Deve (Abast)": "#,##0.00000",
                "Saldo em Casa": "#,##0.00000",
            },
            color_scale_columns=["Deve (Abast)", "Saldo em Casa"],
        )
        if xlsx_ok:
            print(f"[OK] XLSX → {xlsx_path}")


def print_summary(items):
    total_cores = sum(len(it["colors"]) for it in items)
    print(f"\n{'='*65}")
    print(f"  Itens: {len(items)}  |  Cores: {total_cores}")
    print(f"{'='*65}")
    for it in items:
        mf = it.get("mini_fabrica") or "Mini Fabrica - N/D"
        print(f"\n► [{mf}] {it['item_code']} - {it['item_desc']}")
        for c in it["colors"]:
            par_str = f"  [tipo: {c.get('par_tipo')}]" if c.get("par_tipo") else ""
            tam_str = f"  [tam: {c.get('tam')}]" if c.get("tam") else ""
            abast_str = f"  [deve: {c['abast']}]" if c.get("abast") is not None else ""
            casa_str = f"  [em casa: {c['saldo_casa']}]" if c.get("saldo_casa") is not None else ""
            origem_str = f"  [origem: {c.get('saldo_origem', 'nacional')}]"
            print(f"   └── {c['color_code']} - {c['color_desc']}{par_str}{tam_str}{abast_str}{casa_str}{origem_str}")


def build_grouped_items(items, include_singletons=False, include_par_singletons=INCLUDE_PAR_SINGLETONS_IN_COMMON):
    """
    Agrupa por combinação item+tipo+cor+tam.

    include_singletons=False:
      - mantém apenas combinações presentes em 2+ mini fábricas
      - opcionalmente inclui PAR singleton (conforme include_par_singletons)
    include_singletons=True:
      - inclui todas as combinações encontradas
    """
    grouped = {}
    seq = 0
    for it in items:
        mf = it.get("mini_fabrica") or "Mini Fabrica - N/D"
        for c in it.get("colors", []):
            seq += 1
            key = (it.get("item_code"), c.get("par_tipo", ""), c.get("color_code"), c.get("tam", ""))
            if key not in grouped:
                grouped[key] = {
                    "item_code": it.get("item_code"),
                    "item_desc": it.get("item_desc", ""),
                    "par_tipo": c.get("par_tipo", ""),
                    "color_code": c.get("color_code"),
                    "color_desc": c.get("color_desc", ""),
                    "tam": c.get("tam", ""),
                    "por_mini_fabrica": {},
                    "_first_seq": seq,
                    "_mini_fabricas_order": [],
                }
            else:
                # Prefere descrições mais completas quando o mesmo item/cor/tam
                # aparece em múltiplos PDFs com textos truncados.
                current_item_desc = grouped[key].get("item_desc", "") or ""
                new_item_desc = it.get("item_desc", "") or ""
                if len(new_item_desc) > len(current_item_desc):
                    grouped[key]["item_desc"] = new_item_desc

                current_color_desc = grouped[key].get("color_desc", "") or ""
                new_color_desc = c.get("color_desc", "") or ""
                if len(new_color_desc) > len(current_color_desc):
                    grouped[key]["color_desc"] = new_color_desc
            if mf not in grouped[key]["por_mini_fabrica"]:
                grouped[key]["_mini_fabricas_order"].append(mf)
            new_detail = {
                "abast": c.get("abast"),
                "saldo_casa": c.get("saldo_casa"),
                "saldo_origem": c.get("saldo_origem", "nacional"),
                "source_page": c.get("source_page"),
                "source_x": c.get("source_x"),
                "source_y": c.get("source_y"),
                "source_text": c.get("source_text", ""),
            }
            grouped[key]["por_mini_fabrica"][mf] = _merge_destination_detail(
                grouped[key]["por_mini_fabrica"].get(mf),
                new_detail,
            )

    grouped_items = []
    for entry in grouped.values():
        mini_fabs = entry["_mini_fabricas_order"]
        is_par = (entry.get("par_tipo") or "").upper() == "PAR"
        include_single_par = include_par_singletons and is_par and len(mini_fabs) >= 1
        if include_singletons or len(mini_fabs) >= 2 or include_single_par:
            entry["mini_fabricas"] = mini_fabs
            del entry["_mini_fabricas_order"]
            grouped_items.append(entry)
        else:
            del entry["_mini_fabricas_order"]

    grouped_items.sort(key=lambda x: x["_first_seq"])
    for entry in grouped_items:
        del entry["_first_seq"]
    return grouped_items


def build_common_items(items):
    """
    Retorna combinações item+cor presentes em 2+ mini fábricas diferentes.
    """
    return build_grouped_items(
        items,
        include_singletons=False,
        include_par_singletons=INCLUDE_PAR_SINGLETONS_IN_COMMON,
    )


def save_common_json(common_items, path):
    distribution = _build_common_distribution(common_items)
    payload = {
        "total_itens_comuns": len(common_items),
        "total_grupos_distribuicao": len(distribution),
        "grupos_distribuicao": distribution,
    }
    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    print(f"[OK] JSON Comuns → {path}")


def save_common_csv(common_items, path, xlsx_path=None):
    distribution = _build_common_distribution(common_items)
    columns = [
        "Codigo Item", "Descricao Item", "Tipo Unidade", "Codigo Cor", "Descricao Cor", "Tam",
        "Qtd Mini Fabricas", "Mini Fabricas Destino", "Total Necessidade",
        "Necessidade por Mini (deve|saldo|origem)"
    ]
    records = []
    for it in distribution:
        detalhes = " || ".join(
            f"{d['mini_fabrica']}: deve={d.get('deve_abast','')} | saldo={d.get('saldo_casa','')} | origem={d.get('saldo_origem','')}"
            for d in it["destinos"]
        )
        records.append({
            "Codigo Item": it["item_code"],
            "Descricao Item": it["item_desc"],
            "Tipo Unidade": it.get("tipo_unidade", ""),
            "Codigo Cor": it["codigo_cor"],
            "Descricao Cor": it["descricao_cor"],
            "Tam": it.get("tam", ""),
            "Qtd Mini Fabricas": it["qtd_mini_fabricas"],
            "Mini Fabricas Destino": " | ".join(it["mini_fabricas_destino"]),
            "Total Necessidade": it["total_necessidade"],
            "Necessidade por Mini (deve|saldo|origem)": detalhes,
        })

    # Mantém ordem natural dos itens comuns (primeira aparição no fluxo dos PDFs).
    df = pd.DataFrame.from_records(records, columns=columns)
    df = _prepare_dataframe_for_excel(
        df,
        numeric_columns=["Qtd Mini Fabricas", "Total Necessidade"],
    )
    target = Path(path)
    if target.suffix.lower() == ".xlsx":
        xlsx_ok = _save_styled_xlsx_from_df(
            df,
            xlsx_path=str(target),
            sheet_name="ItensComuns",
            table_name="ItensComunsParser",
            numeric_formats={
                "Qtd Mini Fabricas": "0",
                "Total Necessidade": "#,##0.00000",
            },
            color_scale_columns=["Total Necessidade"],
        )
        if xlsx_ok:
            print(f"[OK] XLSX Comuns → {target}")
        else:
            fallback_csv = str(target.with_suffix(".csv"))
            df.to_csv(
                fallback_csv,
                index=False,
                encoding="utf-8-sig",
                sep=";",
                decimal=",",
                quoting=csv.QUOTE_MINIMAL,
            )
            print(f"[WARN] XLSX Comuns indisponível, CSV fallback → {fallback_csv}")
        return

    df.to_csv(path, index=False, encoding="utf-8-sig", sep=";", decimal=",", quoting=csv.QUOTE_MINIMAL)
    print(f"[OK] CSV  Comuns → {path}")
    if xlsx_path:
        xlsx_ok = _save_styled_xlsx_from_df(
            df,
            xlsx_path=xlsx_path,
            sheet_name="ItensComuns",
            table_name="ItensComunsParser",
            numeric_formats={
                "Qtd Mini Fabricas": "0",
                "Total Necessidade": "#,##0.00000",
            },
            color_scale_columns=["Total Necessidade"],
        )
        if xlsx_ok:
            print(f"[OK] XLSX Comuns → {xlsx_path}")


def save_common_html(common_items, html_path, source_label="Múltiplos PDFs", all_items=None):
    distribution_common = _build_common_distribution(common_items)
    if all_items is None:
        distribution_all = distribution_common
    else:
        all_grouped = build_grouped_items(all_items, include_singletons=True, include_par_singletons=True)
        distribution_all = _build_common_distribution(all_grouped)
    data_common_json = json.dumps(distribution_common, ensure_ascii=False)
    data_all_json = json.dumps(distribution_all, ensure_ascii=False)
    generated_at = datetime.now(UTC).strftime("%Y-%m-%d %H:%M:%S UTC")
    html = f"""<!doctype html>
<html lang="pt-BR">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Painel Compacto de Itens em Comum</title>
  <style>
    :root {{
      --bg: #f3f5f7;
      --panel: #fff;
      --line: #dbe2ea;
      --text: #1c2733;
      --muted: #627285;
      --ok-bg: #e8f8f1;
      --ok-tx: #0f766e;
      --warn-bg: #fff7e8;
      --warn-tx: #b45309;
      --bad-bg: #ffe9e9;
      --bad-tx: #b91c1c;
      --tag-bg: #eef5ff;
      --tag-tx: #184e8a;
    }}
    * {{ box-sizing: border-box; }}
    body {{ margin: 0; background: var(--bg); color: var(--text); font: 13px/1.45 "Segoe UI", Tahoma, Arial, sans-serif; }}
    .wrap {{ max-width: 1500px; margin: 0 auto; padding: 12px; }}
    .panel {{ background: var(--panel); border: 1px solid var(--line); border-radius: 10px; padding: 10px; margin-bottom: 10px; }}
    h2, h3 {{ margin: 0; }}
    h2 {{ font-size: 20px; }}
    h3 {{ font-size: 15px; margin-bottom: 8px; }}
    .sub {{ color: var(--muted); margin-top: 4px; }}
    .kpis {{ display: grid; grid-template-columns: repeat(8, minmax(110px, 1fr)); gap: 8px; margin-top: 10px; }}
    .kpi {{ border: 1px solid var(--line); border-radius: 8px; padding: 8px; background: #fbfdff; }}
    .kpi .lbl {{ color: var(--muted); font-size: 11px; text-transform: uppercase; }}
    .kpi .val {{ font-size: 16px; font-weight: 700; margin-top: 2px; }}
    .controls {{ display: grid; grid-template-columns: 2fr 1fr 1fr 1fr 1fr auto auto auto; gap: 8px; align-items: end; }}
    .controls label {{ display: block; font-size: 11px; color: var(--muted); margin-bottom: 3px; text-transform: uppercase; }}
    .controls input, .controls select, .controls button {{ width: 100%; border: 1px solid #c6d3e0; border-radius: 8px; padding: 8px; background: #fff; }}
    .controls button {{ cursor: pointer; font-weight: 600; }}
    .controls .inline {{ display: flex; align-items: center; gap: 6px; color: var(--muted); font-size: 12px; padding-bottom: 2px; }}
    table {{ width: 100%; border-collapse: collapse; }}
    th, td {{ border-bottom: 1px solid #edf1f5; padding: 6px; text-align: left; vertical-align: top; }}
    th {{ background: #f8fbfe; color: #274b6c; position: sticky; top: 0; z-index: 1; font-size: 11px; text-transform: uppercase; letter-spacing: .02em; }}
    .status {{ display: inline-block; border-radius: 999px; padding: 2px 8px; font-size: 11px; font-weight: 700; border: 1px solid transparent; }}
    .s-ok {{ background: var(--ok-bg); color: var(--ok-tx); border-color: #bce9d9; }}
    .s-partial {{ background: var(--warn-bg); color: var(--warn-tx); border-color: #f2d9aa; }}
    .s-none {{ background: var(--bad-bg); color: var(--bad-tx); border-color: #f7c7c7; }}
    .mono {{ font-family: ui-monospace, SFMono-Regular, Menlo, Consolas, monospace; font-size: 12px; }}
    .empty {{ color: var(--muted); padding: 10px; }}
    .small {{ font-size: 12px; color: var(--muted); }}
    .details {{ max-height: 420px; overflow: auto; border: 1px solid var(--line); border-radius: 8px; }}
    .groups {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(420px, 1fr)); gap: 10px; }}
    .group {{ border: 1px solid var(--line); border-radius: 8px; padding: 10px; background: #fcfdff; }}
    .group-top {{ display: flex; align-items: flex-start; justify-content: space-between; gap: 8px; }}
    .group-title {{ font-weight: 700; }}
    .tags {{ display: flex; flex-wrap: wrap; gap: 6px; margin-top: 6px; }}
    .tag {{ background: var(--tag-bg); color: var(--tag-tx); border: 1px solid #d7e7fa; border-radius: 999px; padding: 2px 8px; font-size: 11px; }}
    .mini-list {{ margin-top: 8px; font-size: 12px; }}
    .mini-row {{ display: flex; justify-content: space-between; gap: 8px; border-top: 1px dashed #e6edf5; padding-top: 6px; margin-top: 6px; }}
    .mini-name {{ font-weight: 600; }}
    dialog.side-dialog {{ border: none; padding: 0; max-width: 920px; width: 92vw; margin: 0 0 0 auto; height: 100vh; max-height: 100vh; }}
    dialog.side-dialog::backdrop {{ background: rgba(0, 0, 0, .35); }}
    .dialog-wrap {{ height: 100%; display: grid; grid-template-rows: auto auto 1fr; background: #fff; border-left: 1px solid var(--line); }}
    .dialog-head {{ padding: 12px; border-bottom: 1px solid var(--line); display: flex; justify-content: space-between; align-items: center; }}
    .dialog-actions {{ padding: 10px 12px; border-bottom: 1px solid var(--line); display: flex; gap: 8px; flex-wrap: wrap; align-items: center; }}
    .dialog-actions button {{ border: 1px solid #c6d3e0; border-radius: 8px; background: #fff; padding: 8px 10px; cursor: pointer; font-weight: 600; }}
    .dialog-actions input, .dialog-actions select {{ border: 1px solid #cfd9e5; border-radius: 8px; background: #fff; padding: 8px 10px; min-width: 190px; }}
    .mini-checks {{ display: flex; gap: 10px; flex-wrap: wrap; max-height: 120px; overflow: auto; }}
    .mini-checks label {{ font-size: 12px; color: var(--text); display: inline-flex; align-items: center; gap: 5px; }}
    .dialog-body {{ padding: 12px; overflow: auto; }}
    .warn {{ color: #b45309; font-weight: 600; font-size: 12px; }}
    .ok {{ color: #14532d; font-weight: 700; font-size: 12px; }}
    .compare-summary {{ margin-bottom: 10px; display: flex; flex-wrap: wrap; gap: 6px; }}
    .compare-chip {{ background: #f3f8ff; color: #12416e; border: 1px solid #d8e8f8; border-radius: 999px; padding: 4px 9px; font-size: 12px; }}
    @media (max-width: 1200px) {{ .kpis {{ grid-template-columns: repeat(4, minmax(120px, 1fr)); }} .controls {{ grid-template-columns: 1fr; }} .groups {{ grid-template-columns: 1fr; }} }}
  </style>
</head>
<body>
  <div class="wrap">
    <section class="panel">
      <h2>Itens em Comum - Painel Operacional</h2>
      <div class="sub">Fonte: {escape(source_label)} | Gerado em: {generated_at}</div>
      <div class="kpis">
        <div class="kpi"><div class="lbl">Grupos</div><div class="val" id="kpi-groups">0</div></div>
        <div class="kpi"><div class="lbl">Destinos</div><div class="val" id="kpi-dests">0</div></div>
        <div class="kpi"><div class="lbl">Necessidade</div><div class="val" id="kpi-need">0</div></div>
        <div class="kpi"><div class="lbl">Coberto</div><div class="val" id="kpi-covered">0</div></div>
        <div class="kpi"><div class="lbl">Falta</div><div class="val" id="kpi-falta">0</div></div>
        <div class="kpi"><div class="lbl">Cobertura</div><div class="val" id="kpi-cov">0%</div></div>
        <div class="kpi"><div class="lbl">Sem Cobertura</div><div class="val" id="kpi-none">0</div></div>
        <div class="kpi"><div class="lbl">Saldo Substituto</div><div class="val" id="kpi-sub">0</div></div>
      </div>
    </section>

    <section class="panel controls">
      <div>
        <label for="q">Busca</label>
        <input id="q" placeholder="item, descrição, cor, mini fábrica">
      </div>
      <div>
        <label for="mini">Mini fábrica</label>
        <select id="mini"><option value="">Todas</option></select>
      </div>
      <div>
        <label for="status">Status</label>
        <select id="status">
          <option value="">Todos</option>
          <option value="none">Sem cobertura</option>
          <option value="partial">Parcial</option>
          <option value="ok">Coberto</option>
        </select>
      </div>
      <div>
        <label for="minNeed">Necessidade mínima</label>
        <input id="minNeed" type="number" min="0" step="0.001" value="0">
      </div>
      <div>
        <label for="sort">Ordenação</label>
        <select id="sort">
          <option value="prio">Prioridade</option>
          <option value="falta">Maior falta</option>
          <option value="need">Maior necessidade</option>
          <option value="item">Código item</option>
        </select>
      </div>
      <label class="inline"><input id="onlyFalta" type="checkbox"> Somente com falta</label>
      <label class="inline"><input id="showAll" type="checkbox" checked> Mostrar todos os itens</label>
      <button id="compareMinisBtn" type="button">Comparar Minis</button>
      <button id="exportCsv" type="button">Exportar CSV filtrado</button>
    </section>

    <section class="panel">
      <h3>Roteiro por Destino</h3>
      <div class="small" id="scopeInfo">Carregando...</div>
      <div class="details">
        <table>
          <thead>
            <tr>
              <th>Prio</th>
              <th>Item / Cor</th>
              <th>Destino</th>
              <th>Abrange</th>
              <th>Deve Orig.</th>
              <th>Necess.</th>
              <th>Saldo</th>
              <th>Coberto</th>
              <th>Falta</th>
              <th>Cobertura</th>
              <th>Status</th>
              <th>Origem</th>
            </tr>
          </thead>
          <tbody id="rows"></tbody>
        </table>
      </div>
      <div class="empty" id="empty" style="display:none">Nenhum destino encontrado com os filtros atuais.</div>
    </section>

    <section class="panel">
      <h3>Conferência por Item + Cor (multimini)</h3>
      <div class="small" id="groupsInfo">Carregando...</div>
      <div class="groups" id="groups"></div>
      <div class="empty" id="groupsEmpty" style="display:none">Nenhum grupo encontrado com os filtros atuais.</div>
    </section>
  </div>

  <dialog id="compareDialog" class="side-dialog">
    <div class="dialog-wrap">
      <div class="dialog-head">
        <h3>Comparar Mini Fábricas</h3>
        <button id="compareCloseBtn" type="button">Fechar</button>
      </div>
      <div class="dialog-actions">
        <input id="compareMiniSearch" type="text" placeholder="Filtrar minis no dialog">
        <button id="compareAllBtn" type="button">Marcar visíveis</button>
        <button id="compareNoneBtn" type="button">Limpar seleção</button>
        <select id="compareSort" title="Ordenação">
          <option value="gap">Maior falta</option>
          <option value="covAsc">Menor cobertura</option>
          <option value="need">Maior necessidade</option>
          <option value="mini">Mini (A-Z)</option>
        </select>
        <div class="mini-checks" id="compareMiniList"></div>
        <button id="compareApplyBtn" type="button">Aplicar Comparação</button>
        <span id="compareWarn" class="warn"></span>
        <span id="compareOk" class="ok"></span>
      </div>
      <div class="dialog-body">
        <div id="compareSummary" class="compare-summary"></div>
        <table>
          <thead>
            <tr>
              <th>Mini</th>
              <th>Em comum</th>
              <th>Destinos</th>
              <th>Sem cob.</th>
              <th>Necessidade</th>
              <th>Saldo</th>
              <th>Coberto</th>
              <th>Falta</th>
              <th>Cobertura</th>
            </tr>
          </thead>
          <tbody id="compareRows"></tbody>
        </table>
      </div>
    </div>
  </dialog>

  <script>
    const DATA_COMMON = {data_common_json};
    const DATA_ALL = {data_all_json};
    const rowsEl = document.getElementById('rows');
    const emptyEl = document.getElementById('empty');
    const groupsEl = document.getElementById('groups');
    const groupsEmptyEl = document.getElementById('groupsEmpty');
    const scopeInfo = document.getElementById('scopeInfo');
    const groupsInfo = document.getElementById('groupsInfo');
    const qEl = document.getElementById('q');
    const miniEl = document.getElementById('mini');
    const statusEl = document.getElementById('status');
    const minNeedEl = document.getElementById('minNeed');
    const sortEl = document.getElementById('sort');
    const onlyFaltaEl = document.getElementById('onlyFalta');
    const showAllEl = document.getElementById('showAll');
    const compareMinisBtn = document.getElementById('compareMinisBtn');
    const exportCsvEl = document.getElementById('exportCsv');
    const compareDialog = document.getElementById('compareDialog');
    const compareCloseBtn = document.getElementById('compareCloseBtn');
    const compareApplyBtn = document.getElementById('compareApplyBtn');
    const compareAllBtn = document.getElementById('compareAllBtn');
    const compareNoneBtn = document.getElementById('compareNoneBtn');
    const compareMiniList = document.getElementById('compareMiniList');
    const compareMiniSearch = document.getElementById('compareMiniSearch');
    const compareSort = document.getElementById('compareSort');
    const compareRows = document.getElementById('compareRows');
    const compareWarn = document.getElementById('compareWarn');
    const compareOk = document.getElementById('compareOk');
    const compareSummary = document.getElementById('compareSummary');

    const kpi = {{
      groups: document.getElementById('kpi-groups'),
      dests: document.getElementById('kpi-dests'),
      need: document.getElementById('kpi-need'),
      covered: document.getElementById('kpi-covered'),
      falta: document.getElementById('kpi-falta'),
      cov: document.getElementById('kpi-cov'),
      none: document.getElementById('kpi-none'),
      sub: document.getElementById('kpi-sub')
    }};

    function esc(v) {{
      return String(v ?? '').replace(/[&<>"']/g, c => ({{'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}}[c]));
    }}

    function num(v) {{
      if (v === null || v === undefined || v === '') return 0;
      if (typeof v === 'number') return Number.isFinite(v) ? v : 0;
      const s = String(v).trim();
      const normalized = s.includes(',') && s.includes('.') ? s.replace(/\\./g, '').replace(',', '.') : s.replace(',', '.');
      const n = Number(normalized);
      return Number.isFinite(n) ? n : 0;
    }}

    function fmt(n, d=3) {{ return num(n).toFixed(d); }}

    function statusBy(covered, need) {{
      if (need <= 1e-9) return 'ok';
      if (covered >= need - 1e-9) return 'ok';
      if (covered > 1e-9) return 'partial';
      return 'none';
    }}

    function statusBadge(st) {{
      if (st === 'ok') return '<span class="status s-ok">Coberto</span>';
      if (st === 'partial') return '<span class="status s-partial">Parcial</span>';
      return '<span class="status s-none">Sem cobertura</span>';
    }}

    function sourceTooltip(r) {{
      const page = r.source_page ?? '-';
      const x = r.source_x ?? '-';
      const y = r.source_y ?? '-';
      const raw = String(r.source_text || '').trim();
      return `Página: ${{page}} | x~${{x}} | y~${{y}}\\n${{raw}}`;
    }}

    function flattenRows(groups) {{
      const rows = [];
      for (const g of groups) {{
        const destinos = (g.destinos || []).slice().sort((a,b)=> num(b.necessidade) - num(a.necessidade));
        const groupKey = `${{g.item_code}}|${{g.codigo_cor}}|${{g.tam || ''}}|${{g.tipo_unidade || ''}}`;
        for (const d of destinos) {{
          const need = num(d.necessidade);
          // Cobertura por destino: usa o saldo da própria mini fábrica.
          const saldoDestino = num(d.saldo_casa);
          const covered = Math.min(Math.max(saldoDestino, 0), need);
          const gap = Math.max(need - covered, 0);
          const coveragePct = need > 0 ? (covered / need) * 100 : 100;
          // Prioridade operacional:
          // 1) destinos com falta (>0) vêm primeiro;
          // 2) entre eles, maior falta e maior necessidade sobem;
          // 3) destinos totalmente cobertos vão para o fim.
          const hasGap = gap > 1e-9;
          const prio = hasGap
            ? (gap * 100) + (need * 10) + num(g.qtd_mini_fabricas || 0)
            : (need * 0.01) + (num(g.qtd_mini_fabricas || 0) * 0.001);
          rows.push({{
            group_key: groupKey,
            item_code: g.item_code,
            item_desc: g.item_desc,
            tipo_unidade: g.tipo_unidade || '',
            codigo_cor: g.codigo_cor,
            descricao_cor: g.descricao_cor,
            tam: g.tam || '',
            mini_fabrica: d.mini_fabrica || '',
            deve_abast: d.deve_abast,
            saldo_casa: d.saldo_casa,
            saldo_origem: d.saldo_origem || '',
            source_page: d.source_page,
            source_x: d.source_x,
            source_y: d.source_y,
            source_text: d.source_text || '',
            need,
            covered,
            gap,
            coveragePct,
            status: statusBy(covered, need),
            prio,
            q_mini: num(g.qtd_mini_fabricas || 0)
          }});
        }}
      }}
      return rows;
    }}

    function populateMinis(rows) {{
      const minis = Array.from(new Set(rows.map(r => r.mini_fabrica).filter(Boolean))).sort((a,b)=>a.localeCompare(b));
      miniEl.innerHTML = '<option value="">Todas</option>' + minis.map(m => `<option value="${{esc(m)}}">${{esc(m)}}</option>`).join('');
    }}

    const ALL_ROWS_COMMON = flattenRows(DATA_COMMON);
    const ALL_ROWS_ALL = flattenRows(DATA_ALL);
    let ALL_ROWS = showAllEl.checked ? ALL_ROWS_ALL : ALL_ROWS_COMMON;
    populateMinis(ALL_ROWS);

    function rebuildRowsByMode() {{
      const selectedMini = miniEl.value || '';
      ALL_ROWS = showAllEl.checked ? ALL_ROWS_ALL : ALL_ROWS_COMMON;
      populateMinis(ALL_ROWS);
      if (selectedMini && Array.from(miniEl.options).some(o => o.value === selectedMini)) {{
        miniEl.value = selectedMini;
      }}
      render();
    }}

    function filterBaseRows(ignoreMini = false) {{
      const term = (qEl.value || '').toLowerCase();
      const mini = (miniEl.value || '').toLowerCase();
      const status = statusEl.value || '';
      const minNeed = Math.max(num(minNeedEl.value || 0), 0);
      const onlyFalta = onlyFaltaEl.checked;
      return ALL_ROWS.filter(r => {{
        const hay = [r.item_code, r.item_desc, r.codigo_cor, r.descricao_cor, r.tipo_unidade, r.tam, r.mini_fabrica].join(' ').toLowerCase();
        if (term && !hay.includes(term)) return false;
        if (!ignoreMini && mini && r.mini_fabrica.toLowerCase() !== mini) return false;
        if (status && r.status !== status) return false;
        if (r.need < minNeed) return false;
        if (onlyFalta && r.gap <= 1e-9) return false;
        return true;
      }});
    }}

    function buildCompareMiniOptions() {{
      const minis = Array.from(new Set(ALL_ROWS.map(r => r.mini_fabrica).filter(Boolean))).sort((a,b)=>a.localeCompare(b));
      compareMiniList.innerHTML = minis.map(m => `<label><input type="checkbox" name="compareMini" value="${{esc(m)}}"> ${{esc(m)}}</label>`).join('');
    }}

    function getSelectedCompareMinis() {{
      return Array.from(compareMiniList.querySelectorAll('input[name="compareMini"]:checked')).map(i => i.value);
    }}

    function filterCompareMiniList() {{
      const t = (compareMiniSearch.value || '').toLowerCase().trim();
      const labels = Array.from(compareMiniList.querySelectorAll('label'));
      for (const lb of labels) {{
        const visible = !t || lb.textContent.toLowerCase().includes(t);
        lb.style.display = visible ? 'inline-flex' : 'none';
      }}
    }}

    function renderMiniCompare() {{
      const selected = getSelectedCompareMinis();
      if (selected.length < 2) {{
        compareWarn.textContent = 'Selecione pelo menos 2 mini fábricas.';
        compareOk.textContent = '';
        compareSummary.innerHTML = '';
        compareRows.innerHTML = '';
        return;
      }}
      compareWarn.textContent = '';
      compareOk.textContent = `Comparando ${{selected.length}} minis.`;
      const base = filterBaseRows(true);
      const keyFreq = new Map();
      for (const r of base) {{
        if (!selected.includes(r.mini_fabrica)) continue;
        const k = `${{r.mini_fabrica}}|${{r.group_key}}`;
        if (!keyFreq.has(k)) keyFreq.set(k, true);
      }}
      const groupCountByMini = new Map();
      const globalGroupFreq = new Map();
      for (const k of keyFreq.keys()) {{
        const [mini, gk] = k.split('|');
        if (!groupCountByMini.has(mini)) groupCountByMini.set(mini, new Set());
        groupCountByMini.get(mini).add(gk);
        globalGroupFreq.set(gk, (globalGroupFreq.get(gk) || 0) + 1);
      }}
      const out = [];
      for (const mini of selected) {{
        const rows = base.filter(r => r.mini_fabrica === mini);
        const need = rows.reduce((acc, r) => acc + r.need, 0);
        const saldo = rows.reduce((acc, r) => acc + num(r.saldo_casa), 0);
        const covered = rows.reduce((acc, r) => acc + r.covered, 0);
        const gap = rows.reduce((acc, r) => acc + r.gap, 0);
        const cov = need > 0 ? (covered / need) * 100 : 100;
        const noneCount = rows.filter(r => r.status === 'none').length;
        const miniGroups = groupCountByMini.get(mini) || new Set();
        let commonGroups = 0;
        for (const gk of miniGroups) {{
          if ((globalGroupFreq.get(gk) || 0) > 1) commonGroups += 1;
        }}
        out.push({{ mini, commonGroups, destinos: rows.length, noneCount, need, saldo, covered, gap, cov }});
      }}
      if (compareSort.value === 'mini') out.sort((a,b)=> a.mini.localeCompare(b.mini));
      else if (compareSort.value === 'need') out.sort((a,b)=> b.need - a.need);
      else if (compareSort.value === 'covAsc') out.sort((a,b)=> a.cov - b.cov || b.gap - a.gap);
      else out.sort((a,b)=> b.gap - a.gap || b.need - a.need);

      const totalNeed = out.reduce((a,r)=>a+r.need, 0);
      const totalGap = out.reduce((a,r)=>a+r.gap, 0);
      const totalCovered = out.reduce((a,r)=>a+r.covered, 0);
      const totalCov = totalNeed > 0 ? (totalCovered / totalNeed) * 100 : 100;
      compareSummary.innerHTML = [
        `<span class="compare-chip">Necessidade total: ${{fmt(totalNeed, 1)}}</span>`,
        `<span class="compare-chip">Falta total: ${{fmt(totalGap, 1)}}</span>`,
        `<span class="compare-chip">Cobertura média: ${{totalCov.toFixed(1)}}%</span>`
      ].join('');
      compareRows.innerHTML = out.map(r => `
        <tr>
          <td><strong>${{esc(r.mini)}}</strong></td>
          <td class="mono">${{r.commonGroups}}</td>
          <td class="mono">${{r.destinos}}</td>
          <td class="mono">${{r.noneCount}}</td>
          <td class="mono">${{fmt(r.need, 1)}}</td>
          <td class="mono">${{fmt(r.saldo, 1)}}</td>
          <td class="mono">${{fmt(r.covered, 1)}}</td>
          <td class="mono">${{fmt(r.gap, 1)}}</td>
          <td class="mono">${{r.cov.toFixed(1)}}%</td>
        </tr>
      `).join('');
    }}

    function applyFilters() {{
      const term = (qEl.value || '').toLowerCase();
      const mini = (miniEl.value || '').toLowerCase();
      const status = statusEl.value || '';
      const minNeed = Math.max(num(minNeedEl.value || 0), 0);
      const onlyFalta = onlyFaltaEl.checked;

      let rows = ALL_ROWS.filter(r => {{
        const hay = [r.item_code, r.item_desc, r.codigo_cor, r.descricao_cor, r.tipo_unidade, r.tam, r.mini_fabrica].join(' ').toLowerCase();
        if (term && !hay.includes(term)) return false;
        if (mini && r.mini_fabrica.toLowerCase() !== mini) return false;
        if (status && r.status !== status) return false;
        if (r.need < minNeed) return false;
        if (onlyFalta && r.gap <= 1e-9) return false;
        return true;
      }});

      if (sortEl.value === 'item') rows.sort((a,b)=> String(a.item_code).localeCompare(String(b.item_code)));
      else if (sortEl.value === 'need') rows.sort((a,b)=> b.need - a.need);
      else if (sortEl.value === 'falta') rows.sort((a,b)=> b.gap - a.gap);
      else rows.sort((a,b)=> b.prio - a.prio);

      return rows;
    }}

    function summarizeGroups(rows) {{
      const map = new Map();
      for (const r of rows) {{
        if (!map.has(r.group_key)) {{
          map.set(r.group_key, {{
            group_key: r.group_key,
            item_code: r.item_code,
            item_desc: r.item_desc,
            codigo_cor: r.codigo_cor,
            descricao_cor: r.descricao_cor,
            tipo_unidade: r.tipo_unidade,
            tam: r.tam,
            q_mini: r.q_mini,
            need: 0,
            covered: 0,
            gap: 0,
            prio: 0,
            has_sub: false,
            destinos: []
          }});
        }}
        const g = map.get(r.group_key);
        g.need += r.need;
        g.covered += r.covered;
        g.gap += r.gap;
        g.prio += r.prio;
        g.has_sub = g.has_sub || String(r.saldo_origem || '').toLowerCase() === 'substituto';
        g.destinos.push(r);
      }}

      const groups = Array.from(map.values()).map(g => {{
        g.coveragePct = g.need > 0 ? (g.covered / g.need) * 100 : 100;
        g.status = statusBy(g.covered, g.need);
        g.destinos.sort((a,b)=> b.gap - a.gap);
        return g;
      }});

      if (sortEl.value === 'item') groups.sort((a,b)=> String(a.item_code).localeCompare(String(b.item_code)));
      else if (sortEl.value === 'need') groups.sort((a,b)=> b.need - a.need);
      else if (sortEl.value === 'falta') groups.sort((a,b)=> b.gap - a.gap);
      else groups.sort((a,b)=> b.prio - a.prio);
      return groups;
    }}

    function render() {{
      const rows = applyFilters();
      const groups = summarizeGroups(rows);
      const modeLabel = showAllEl.checked ? "todos os itens" : "somente itens em comum";
      const totalNeed = rows.reduce((acc, r) => acc + r.need, 0);
      const totalCovered = rows.reduce((acc, r) => acc + r.covered, 0);
      const totalGap = Math.max(totalNeed - totalCovered, 0);
      const covPct = totalNeed > 0 ? (totalCovered / totalNeed) * 100 : 100;

      kpi.groups.textContent = String(groups.length);
      kpi.dests.textContent = String(rows.length);
      kpi.need.textContent = fmt(totalNeed, 1);
      kpi.covered.textContent = fmt(totalCovered, 1);
      kpi.falta.textContent = fmt(totalGap, 1);
      kpi.cov.textContent = `${{covPct.toFixed(1)}}%`;
      kpi.none.textContent = String(rows.filter(r => r.status === 'none').length);
      kpi.sub.textContent = String(rows.filter(r => String(r.saldo_origem || '').toLowerCase() === 'substituto').length);
      scopeInfo.textContent = `Visão filtrada (${{modeLabel}}): ${{rows.length}} destino(s), ${{groups.length}} grupo(s), falta total ${{fmt(totalGap,1)}}.`;
      groupsInfo.textContent = `Conferência por item+cor para ${{modeLabel}} no recorte atual.`;

      if (!rows.length) {{
        rowsEl.innerHTML = '';
        emptyEl.style.display = 'block';
      }} else {{
        emptyEl.style.display = 'none';
        rowsEl.innerHTML = rows.map(r => `
          <tr>
            <td class="mono">${{r.prio.toFixed(1)}}</td>
            <td>
              <div><strong>${{esc(r.item_code)}} - ${{esc(r.item_desc)}}</strong></div>
              <div class="small">Cor: ${{esc(r.codigo_cor)}} - ${{esc(r.descricao_cor)}} | Tipo: ${{esc(r.tipo_unidade || '-')}} | Tam: ${{esc(r.tam || '-')}}</div>
              <div class="small"><span class="tag" title="${{esc(sourceTooltip(r))}}">Origem PDF</span></div>
            </td>
            <td><strong>${{esc(r.mini_fabrica)}}</strong></td>
            <td class="mono">${{r.q_mini}}</td>
            <td class="mono">${{fmt(r.deve_abast)}}</td>
            <td class="mono">${{fmt(r.need)}}</td>
            <td class="mono">${{fmt(r.saldo_casa)}}</td>
            <td class="mono">${{fmt(r.covered)}}</td>
            <td class="mono">${{fmt(r.gap)}}</td>
            <td class="mono">${{r.coveragePct.toFixed(0)}}%</td>
            <td>${{statusBadge(r.status)}}</td>
            <td>${{esc(r.saldo_origem || '-')}}</td>
          </tr>
        `).join('');
      }}

      if (!groups.length) {{
        groupsEl.innerHTML = '';
        groupsEmptyEl.style.display = 'block';
      }} else {{
        groupsEmptyEl.style.display = 'none';
        groupsEl.innerHTML = groups.map(g => `
          <article class="group">
            <div class="group-top">
              <div>
                <div class="group-title">${{esc(g.item_code)}} - ${{esc(g.item_desc)}}</div>
                <div class="small">Cor: ${{esc(g.codigo_cor)}} - ${{esc(g.descricao_cor)}} | Tipo: ${{esc(g.tipo_unidade || '-')}} | Tam: ${{esc(g.tam || '-')}}</div>
              </div>
              <div>${{statusBadge(g.status)}}</div>
            </div>
            <div class="tags">
              <span class="tag">Minis: ${{g.q_mini}}</span>
              <span class="tag">Necessidade: ${{fmt(g.need, 1)}}</span>
              <span class="tag">Falta: ${{fmt(g.gap, 1)}}</span>
              <span class="tag">Cobertura: ${{g.coveragePct.toFixed(1)}}%</span>
              <span class="tag">Prio: ${{g.prio.toFixed(1)}}</span>
              ${{g.has_sub ? '<span class="tag">Tem saldo de substituto</span>' : ''}}
            </div>
            <div class="mini-list">
              ${{g.destinos.map(d => `
                <div class="mini-row">
                  <div class="mini-name">${{esc(d.mini_fabrica)}}</div>
                  <div class="mono">deve ${{fmt(d.deve_abast)}} | nec ${{fmt(d.need)}} | falta ${{fmt(d.gap)}} | cob ${{d.coveragePct.toFixed(0)}}%</div>
                </div>
              `).join('')}}
            </div>
          </article>
        `).join('');
      }}
    }}

    [qEl, miniEl, statusEl, minNeedEl, sortEl, onlyFaltaEl].forEach(el => {{
      const evt = el.tagName === 'INPUT' && el.type !== 'checkbox' ? 'input' : 'change';
      el.addEventListener(evt, render);
    }});
    showAllEl.addEventListener('change', rebuildRowsByMode);
    compareMinisBtn.addEventListener('click', () => {{
      buildCompareMiniOptions();
      compareWarn.textContent = '';
      compareOk.textContent = '';
      compareMiniSearch.value = '';
      filterCompareMiniList();
      compareSummary.innerHTML = '';
      compareRows.innerHTML = '';
      compareDialog.showModal();
    }});
    compareCloseBtn.addEventListener('click', () => compareDialog.close());
    compareApplyBtn.addEventListener('click', renderMiniCompare);
    compareMiniSearch.addEventListener('input', filterCompareMiniList);
    compareSort.addEventListener('change', renderMiniCompare);
    compareAllBtn.addEventListener('click', () => {{
      for (const cb of compareMiniList.querySelectorAll('input[name="compareMini"]')) {{
        const label = cb.closest('label');
        if (label && label.style.display === 'none') continue;
        cb.checked = true;
      }}
    }});
    compareNoneBtn.addEventListener('click', () => {{
      for (const cb of compareMiniList.querySelectorAll('input[name="compareMini"]')) cb.checked = false;
      compareSummary.innerHTML = '';
      compareRows.innerHTML = '';
      compareWarn.textContent = '';
      compareOk.textContent = '';
    }});

    function exportFilteredCsv() {{
      const rows = applyFilters();
        const header = [
          "prioridade","codigo_item","descricao_item","tipo_unidade","codigo_cor","descricao_cor","tam",
          "mini_fabrica_destino","qtd_mini_fabricas","deve_original","necessidade","saldo_casa","coberto","falta","cobertura_pct","status","origem_saldo"
        ];
      const csvRows = [header];
      for (const r of rows) {{
        csvRows.push([
          r.prio.toFixed(1),
          r.item_code,
          r.item_desc,
          r.tipo_unidade || "",
          r.codigo_cor,
          r.descricao_cor,
          r.tam || "",
          r.mini_fabrica,
          r.q_mini,
          fmt(r.deve_abast),
          fmt(r.need),
          fmt(r.saldo_casa),
          fmt(r.covered),
          fmt(r.gap),
          r.coveragePct.toFixed(2),
          r.status,
          r.saldo_origem || ""
        ]);
      }}
      const escCell = (v) => `"${{String(v ?? "").replace(/"/g, '""')}}"`;
      const csvText = csvRows.map(row => row.map(escCell).join(",")).join("\\n");
      const blob = new Blob([csvText], {{ type: "text/csv;charset=utf-8;" }});
      const url = URL.createObjectURL(blob);
      const a = document.createElement("a");
      a.href = url;
      a.download = "common_items_filtrado.csv";
      document.body.appendChild(a);
      a.click();
      a.remove();
      URL.revokeObjectURL(url);
    }}
    exportCsvEl.addEventListener("click", exportFilteredCsv);
    render();
  </script>
</body>
</html>"""
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"[OK] HTML Comuns → {html_path}")
def print_common_summary(common_items):
    distribution = _build_common_distribution(common_items)
    print(f"\n{'='*65}")
    print(f"  Itens em Comum (mesmo item + mesma cor): {len(common_items)}")
    print(f"{'='*65}")
    for it in distribution:
        mfs = ", ".join(it["mini_fabricas_destino"])
        print(f"\n► {it['item_code']} - {it['item_desc']} | {it['codigo_cor']} - {it['descricao_cor']}")
        print(f"   └── Destinos ({it['qtd_mini_fabricas']}): {mfs} | necessidade total: {it['total_necessidade']}")


if FASTAPI_AVAILABLE:
    TAGS_METADATA = [
        {
            "name": "status",
            "description": "Saude e disponibilidade da API.",
        },
        {
            "name": "files",
            "description": "Navegacao e download dos arquivos gerados em JSON/CSV/HTML.",
        },
        {
            "name": "parser",
            "description": (
                "Processamento de PDF(s) com extracao de itens/cores, mini fabrica, "
                "saldo em casa, origem do saldo, tipo PAR e Tam, alem da distribuicao de itens comuns."
            ),
        },
    ]

    class HealthResponse(BaseModel):
        status: str = Field(description="Status da API.", examples=["ok"])

    class GeneratedFile(BaseModel):
        name: str = Field(description="Nome do arquivo.")
        relative_path: str = Field(description="Caminho relativo dentro da pasta de output.")
        size_bytes: int = Field(description="Tamanho em bytes.")
        updated_at: str = Field(description="Timestamp UTC ISO-8601 de atualizacao.")
        direct_url: str = Field(
            description="Link direto para download/abertura do arquivo via endpoint /files/{filename}."
        )

    class FileListResponse(BaseModel):
        output_dir: str
        count: int
        files: List[GeneratedFile]

    class ParseColor(BaseModel):
        color_code: str
        color_desc: str
        par_tipo: str = Field(default="", description="Tipo de unidade, ex.: PAR.")
        tam: str = Field(default="", description="Tamanho (Tam), quando aplicavel.")
        abast: Optional[float] = None
        saldo_casa: Optional[str] = None
        saldo_origem: str = Field(default="nacional")

    class ParseItem(BaseModel):
        item_code: str
        item_desc: str
        mini_fabrica: Optional[str] = None
        colors: List[ParseColor]

    class ParseSummary(BaseModel):
        total_items: int
        total_colors: int
        total_itens_comuns: int = 0

    class ParseOutputs(BaseModel):
        json: str = Field(description="Link direto do JSON principal.")
        csv: str = Field(description="Link direto do CSV principal.")
        html: str = Field(description="Link direto do HTML principal.")
        json_url: str = Field(description="Alias explícito para o link direto do JSON principal.")
        csv_url: str = Field(description="Alias explícito para o link direto do CSV principal.")
        html_url: str = Field(description="Alias explícito para o link direto do HTML principal.")
        json_local_path: str = Field(description="Caminho local do JSON principal no servidor.")
        csv_local_path: str = Field(description="Caminho local do CSV principal no servidor.")
        html_local_path: str = Field(description="Caminho local do HTML principal no servidor.")
        common_json: Optional[str] = Field(default=None, description="Link direto do JSON de itens em comum.")
        common_csv: Optional[str] = Field(default=None, description="Link direto do CSV de itens em comum.")
        common_html: Optional[str] = Field(default=None, description="Link direto do HTML de itens em comum.")
        common_json_url: Optional[str] = Field(
            default=None, description="Alias explícito para o link direto do JSON de itens em comum."
        )
        common_csv_url: Optional[str] = Field(
            default=None, description="Alias explícito para o link direto do CSV de itens em comum."
        )
        common_html_url: Optional[str] = Field(
            default=None, description="Alias explícito para o link direto do HTML de itens em comum."
        )

    class ParseResponse(BaseModel):
        files: List[str]
        summary: ParseSummary
        outputs: ParseOutputs
        items: List[ParseItem]

    app = FastAPI(
        title="PDF Parser API",
        version="2.0.0",
        summary="Parser de saldo de abastecimento (single e multi-PDF).",
        description=(
            "API para processar relatórios PDF (fab0257) e gerar saídas estruturadas.\n\n"
            "Funcionalidades atuais:\n"
            "- Upload de 1 ou vários PDFs no mesmo request.\n"
            "- Extração de item, cor, mini fábrica, abast (deve), saldo em casa e origem do saldo.\n"
            "- Suporte a linhas de cor no formato numérico e no formato PAR (ex.: PAR107848).\n"
            "- Extração de Tam (tamanho) para linhas PAR.\n"
            "- Fallback de saldo do substituto quando saldo do nacional é zero.\n"
            "- Geração de JSON, CSV e HTML para parse principal.\n"
            "- Geração de JSON, CSV e HTML de itens comuns (distribuição entre mini fábricas).\n"
            "- Listagem e download dos arquivos gerados via API."
        ),
        openapi_tags=TAGS_METADATA,
        contact={"name": "Parser API Support", "url": "https://github.com/vinicius/n8n-manager"},
        license_info={"name": "MIT", "identifier": "MIT"},
    )

    def custom_openapi():
        if app.openapi_schema:
            return app.openapi_schema
        openapi_schema = get_openapi(
            title=app.title,
            version=app.version,
            summary=app.summary,
            description=app.description,
            routes=app.routes,
            tags=app.openapi_tags,
            contact=app.contact,
            license_info=app.license_info,
        )
        openapi_schema["externalDocs"] = {
            "description": "Repositório do projeto",
            "url": "https://github.com/vinicius/n8n-manager",
        }
        app.openapi_schema = openapi_schema
        return app.openapi_schema

    app.openapi = custom_openapi

    def _output_dir() -> Path:
        output_dir = Path(os.getenv("PARSER_OUTPUT_DIR", "parsed_output"))
        output_dir.mkdir(parents=True, exist_ok=True)
        return output_dir

    def _public_base_url(request: Request) -> str:
        forwarded_proto = request.headers.get("x-forwarded-proto")
        forwarded_host = request.headers.get("x-forwarded-host") or request.headers.get("host")
        if forwarded_proto and forwarded_host:
            return f"{forwarded_proto}://{forwarded_host}".rstrip("/")
        return str(request.base_url).rstrip("/")

    def _safe_file_path(filename: str) -> Path:
        output_dir = _output_dir().resolve()
        file_path = (output_dir / filename).resolve()
        if output_dir not in file_path.parents and file_path != output_dir:
            raise HTTPException(status_code=400, detail="Arquivo invalido")
        if not file_path.is_file():
            raise HTTPException(status_code=404, detail="Arquivo nao encontrado")
        return file_path

    @app.get(
        "/health",
        tags=["status"],
        summary="Health check da API",
        description="Retorna estado de disponibilidade da API.",
        response_model=HealthResponse,
    )
    def health():
        return {"status": "ok"}

    @app.get(
        "/files",
        tags=["files"],
        summary="Lista arquivos gerados",
        description=(
            "Lista todos os arquivos presentes na pasta de output da parser API. "
            "Inclui nome, caminho relativo, tamanho e data de atualização."
        ),
        response_model=FileListResponse,
    )
    def list_files(request: Request):
        output_dir = _output_dir().resolve()
        base_url = _public_base_url(request)
        files = []
        for path in sorted(output_dir.rglob("*")):
            if not path.is_file():
                continue
            rel = path.relative_to(output_dir).as_posix()
            files.append(
                {
                    "name": path.name,
                    "relative_path": rel,
                    "size_bytes": path.stat().st_size,
                    "updated_at": datetime.fromtimestamp(path.stat().st_mtime, UTC).isoformat().replace("+00:00", "Z"),
                    "direct_url": f"{base_url}/files/{quote(rel)}",
                }
            )
        return {"output_dir": str(output_dir), "count": len(files), "files": files}

    @app.get(
        "/files/{filename:path}",
        tags=["files"],
        summary="Download de arquivo gerado",
        description=(
            "Faz download de um arquivo gerado pela API (JSON, CSV, HTML ou binário). "
            "O caminho deve ser relativo à pasta de output."
        ),
        responses={
            200: {"description": "Arquivo retornado com sucesso."},
            400: {"description": "Arquivo inválido."},
            404: {"description": "Arquivo não encontrado."},
        },
    )
    def download_file(filename: str):
        file_path = _safe_file_path(filename)
        suffix = file_path.suffix.lower()
        media_type = "application/octet-stream"
        if suffix == ".json":
            media_type = "application/json"
        elif suffix == ".csv":
            media_type = "text/csv"
        elif suffix == ".xlsx":
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        elif suffix == ".html":
            media_type = "text/html"
        return FileResponse(str(file_path), filename=file_path.name, media_type=media_type)

    @app.post(
        "/parse",
        tags=["parser"],
        summary="Processa 1 ou vários PDFs",
        description=(
            "Recebe arquivo(s) PDF e gera saídas estruturadas.\n\n"
            "Entradas aceitas:\n"
            "- `file`: envio de 1 PDF\n"
            "- `files`: envio de múltiplos PDFs no mesmo request\n\n"
            "Saídas geradas:\n"
            "- Parse principal: JSON e CSV\n"
            "- Itens comuns (distribuição): JSON, CSV e HTML (painel único)\n\n"
            "Links diretos:\n"
            "- O campo `outputs` retorna links diretos de todos os arquivos gerados (`json/csv/html` e `common_*`).\n"
            "- O endpoint `/files` lista cada arquivo com `direct_url` pronto para download/abertura.\n\n"
            "Regras aplicadas no parse:\n"
            "- Mini fábrica por contexto de cabeçalho do relatório\n"
            "- Cor numérica e cor PAR (com `Tipo Unidade=PAR` e `Tam`)\n"
            "- `saldo_casa` com fallback para substituto quando saldo nacional é zero"
        ),
        response_model=ParseResponse,
        responses={
            200: {"description": "PDF(s) processado(s) com sucesso."},
            400: {"description": "Arquivo ausente ou extensão inválida."},
            500: {"description": "Falha durante o processamento."},
        },
    )
    async def parse(
        request: Request,
        files: Optional[List[UploadFile]] = File(default=None),
        file: Optional[UploadFile] = File(default=None),
    ):
        upload_files: List[UploadFile] = []
        if files:
            upload_files.extend(files)
        if file is not None:
            upload_files.append(file)

        if not upload_files:
            raise HTTPException(status_code=400, detail="Envie pelo menos um arquivo PDF.")
        if any(not (f.filename or "").lower().endswith(".pdf") for f in upload_files):
            raise HTTPException(status_code=400, detail="Todos os arquivos devem ser PDF.")

        temp_paths: List[Path] = []
        output_dir = _output_dir()
        base_url = _public_base_url(request)

        try:
            all_items = []
            uploaded_names = []
            for up in upload_files:
                uploaded_names.append(up.filename or "arquivo.pdf")
                suffix = Path(up.filename or "arquivo.pdf").suffix or ".pdf"
                with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
                    tmp.write(await up.read())
                    temp_path = Path(tmp.name)
                temp_paths.append(temp_path)
                all_items.extend(parse_pdf(str(temp_path)))

            stamp = datetime.now(UTC).strftime("%Y%m%d%H%M%S")
            spreadsheet_ext = "xlsx" if OPENPYXL_AVAILABLE else "csv"
            if not OPENPYXL_AVAILABLE:
                print("[WARN] openpyxl não instalado: exportação mantida em CSV.")
            if len(upload_files) == 1:
                base_name = Path(upload_files[0].filename or "arquivo").stem
                json_name = f"{base_name}_{stamp}_parsed.json"
                csv_name = f"{base_name}_{stamp}_parsed.{spreadsheet_ext}"
                common_json_name = f"{base_name}_{stamp}_common_items.json"
                common_csv_name = f"{base_name}_{stamp}_common_items.{spreadsheet_ext}"
                common_html_name = f"{base_name}_{stamp}_common_items.html"
                html_name = common_html_name
                json_path = output_dir / json_name
                csv_path = output_dir / csv_name
                html_path = output_dir / html_name
                common_json_path = output_dir / common_json_name
                common_csv_path = output_dir / common_csv_name
                common_html_path = output_dir / common_html_name
                save_json(all_items, str(json_path))
                save_csv(all_items, str(csv_path))
                common_items = build_common_items(all_items)
                save_common_json(common_items, str(common_json_path))
                save_common_csv(common_items, str(common_csv_path))
                save_common_html(
                    common_items,
                    str(common_html_path),
                    source_label=uploaded_names[0],
                    all_items=all_items,
                )
            else:
                json_name = f"multi_pdf_{stamp}_parsed.json"
                csv_name = f"multi_pdf_{stamp}_parsed.{spreadsheet_ext}"
                common_json_name = f"multi_pdf_{stamp}_common_items.json"
                common_csv_name = f"multi_pdf_{stamp}_common_items.{spreadsheet_ext}"
                common_html_name = f"multi_pdf_{stamp}_common_items.html"
                html_name = common_html_name
                json_path = output_dir / json_name
                csv_path = output_dir / csv_name
                html_path = output_dir / html_name
                common_json_path = output_dir / common_json_name
                common_csv_path = output_dir / common_csv_name
                common_html_path = output_dir / common_html_name
                save_json(all_items, str(json_path))
                save_csv(all_items, str(csv_path))
                common_items = build_common_items(all_items)
                save_common_json(common_items, str(common_json_path))
                save_common_csv(common_items, str(common_csv_path))
                save_common_html(
                    common_items,
                    str(common_html_path),
                    source_label=f"{len(uploaded_names)} PDFs",
                    all_items=all_items,
                )

            total_colors = sum(len(it["colors"]) for it in all_items)
            response = {
                "files": uploaded_names,
                "summary": {"total_items": len(all_items), "total_colors": total_colors},
                "outputs": {
                    "json": f"{base_url}/files/{quote(json_name)}",
                    "csv": f"{base_url}/files/{quote(csv_name)}",
                    "html": f"{base_url}/files/{quote(html_name)}",
                    "json_url": f"{base_url}/files/{quote(json_name)}",
                    "csv_url": f"{base_url}/files/{quote(csv_name)}",
                    "html_url": f"{base_url}/files/{quote(html_name)}",
                    "json_local_path": str(json_path),
                    "csv_local_path": str(csv_path),
                    "html_local_path": str(html_path),
                },
                "items": all_items,
            }
            response["summary"]["total_itens_comuns"] = len(common_items)
            response["outputs"]["common_json"] = f"{base_url}/files/{quote(common_json_name)}"
            response["outputs"]["common_csv"] = f"{base_url}/files/{quote(common_csv_name)}"
            response["outputs"]["common_html"] = f"{base_url}/files/{quote(common_html_name)}"
            response["outputs"]["common_json_url"] = f"{base_url}/files/{quote(common_json_name)}"
            response["outputs"]["common_csv_url"] = f"{base_url}/files/{quote(common_csv_name)}"
            response["outputs"]["common_html_url"] = f"{base_url}/files/{quote(common_html_name)}"

            return response
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Falha ao processar PDF(s): {e}")
        finally:
            for temp_path in temp_paths:
                try:
                    temp_path.unlink(missing_ok=True)
                except OSError:
                    pass
else:
    app = None


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python parse_falta.py <caminho_do_pdf1> [caminho_do_pdf2 ...]")
        sys.exit(1)

    pdf_paths = sys.argv[1:]

    spreadsheet_ext = "xlsx" if OPENPYXL_AVAILABLE else "csv"

    # Modo único: 1 PDF -> saídas parse (JSON/planilha) + comuns (JSON/planilha/HTML).
    if len(pdf_paths) == 1:
        pdf_path = pdf_paths[0]
        print(f"Processando: {pdf_path}\n")
        items = parse_pdf(pdf_path)

        base = os.path.splitext(pdf_path)[0]
        save_json(items, base + "_parsed.json")
        save_csv(items,  base + f"_parsed.{spreadsheet_ext}")
        print_summary(items)
        common_items = build_common_items(items)
        save_common_json(common_items, base + "_common_items.json")
        save_common_csv(common_items, base + f"_common_items.{spreadsheet_ext}")
        save_common_html(
            common_items,
            base + "_common_items.html",
            source_label=os.path.basename(pdf_path),
            all_items=items,
        )
        print_common_summary(common_items)
    else:
        # Modo consolidado: 2+ PDFs -> 1 JSON/planilha parse + 1 JSON/planilha/HTML comuns.
        all_items = []
        for idx, pdf_path in enumerate(pdf_paths, start=1):
            if idx > 1:
                print("\n" + "#" * 80 + "\n")
            print(f"Processando ({idx}/{len(pdf_paths)}): {pdf_path}")
            all_items.extend(parse_pdf(pdf_path))

        first_dir = os.path.dirname(os.path.abspath(pdf_paths[0]))
        out_json = os.path.join(first_dir, "multi_pdf_parsed.json")
        out_csv = os.path.join(first_dir, f"multi_pdf_parsed.{spreadsheet_ext}")
        out_common_json = os.path.join(first_dir, "multi_pdf_common_items.json")
        out_common_csv = os.path.join(first_dir, f"multi_pdf_common_items.{spreadsheet_ext}")
        out_common_html = os.path.join(first_dir, "multi_pdf_common_items.html")

        print(f"\nGerando saída consolidada de {len(pdf_paths)} PDFs...\n")
        save_json(all_items, out_json)
        save_csv(all_items, out_csv)
        print_summary(all_items)

        common_items = build_common_items(all_items)
        save_common_json(common_items, out_common_json)
        save_common_csv(common_items, out_common_csv)
        save_common_html(
            common_items,
            out_common_html,
            source_label=f"{len(pdf_paths)} PDFs",
            all_items=all_items,
        )
        print_common_summary(common_items)
